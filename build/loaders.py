import re
import openpyxl

_WEEK_RE = re.compile(r"nfl week\s*(\d+)", re.I)
_PLAYOFF_RE = re.compile(r"playoff round\s*(\d+).*week\s*(\d+)", re.I)


def _isnum(v):
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def parse_league_history(path):
    """Parse every season sheet into a flat list of raw game dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    games = []
    for name in wb.sheetnames:
        if not name.strip().isdigit():
            continue
        season = int(name.strip())
        ws = wb[name]
        phase = "regular"
        week = None
        rnd = None
        for row in ws.iter_rows(values_only=True):
            c0 = row[0]
            if isinstance(c0, str):
                low = c0.strip().lower()
                mp = _PLAYOFF_RE.search(low)
                if mp:
                    phase, rnd, week = "playoff", int(mp.group(1)), int(mp.group(2))
                    continue
                mw = _WEEK_RE.search(low)
                if mw:
                    phase, rnd, week = "regular", None, int(mw.group(1))
                    continue
                if low.startswith("away team") or low.startswith("bye"):
                    continue
            # data row: cols = away_team, away_mgr, away_score, home_score, home_mgr, home_team
            if len(row) >= 6 and _isnum(row[2]) and _isnum(row[3]):
                games.append({
                    "season": season,
                    "week": week,
                    "phase": phase,
                    "playoff_round": rnd,
                    "away_team_raw": str(row[0]).strip() if row[0] else "",
                    "away_manager": str(row[1]).strip() if row[1] else "",
                    "away_score": round(float(row[2]), 1),
                    "home_score": round(float(row[3]), 1),
                    "home_manager": str(row[4]).strip() if row[4] else "",
                    "home_team_raw": str(row[5]).strip() if row[5] else "",
                })
    wb.close()
    return games


def parse_gridiron(path):
    """Read Data_Sorted into {(season, week, frozenset(owners)): {owner: score}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Data_Sorted"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # cols: A blank, B season, C week, D home, E home_score, F away, G away_score
        season, week, home, hs, away, as_ = row[1], row[2], row[3], row[4], row[5], row[6]
        if not (_isnum(season) and _isnum(hs) and _isnum(as_)):
            continue
        if not home or not away:
            continue
        season = int(float(season))
        wk = int(float(week)) if _isnum(week) else week
        out[(season, wk, frozenset({home, away}))] = {
            home: round(float(hs), 1),
            away: round(float(as_), 1),
        }
    wb.close()
    return out
